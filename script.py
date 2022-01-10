import requests, csv
from bs4 import BeautifulSoup
from datetime import date
from openpyxl import Workbook

# Constant(s)
today = date.today()
cur_month = today.month
if cur_month == 1:
    month_ago = today.replace(month=12)
else:
    month_ago = today.replace(month=(cur_month - 1))

# Reformat Dates for hist data URL Parameters
today = today.strftime("%m/%d/%Y")
month_ago = month_ago.strftime("%m/%d/%Y")

# Define the stock symbols that we want to use, eventually with a separate file
fn = "symbols.csv"
wb = Workbook()
ws = wb.active
wb.remove(ws)

with open(fn) as f:
    lines = csv.reader(f)
    for sym in lines:
        if sym == "":
            break
        elif sym[0] == "SYMBOLS":
            pass

        # Define stock symbol e.g. "AAPL"
        symbol = str(sym[0])

        # Get basic Overview data
        o_url = str(f"https://finance.yahoo.com/quote/{symbol}?p={symbol}")
        r = requests.get(o_url)
        soup = BeautifulSoup(r.content, "html.parser")

        # Define variables for Overview
        results = soup.find(id="quote-summary")
        corp = (soup.find("h1", class_="D(ib) Fz(18px)")).text.strip()
        prev_close = (
            results.find("td", attrs={"data-test": "PREV_CLOSE-value"})
        ).text.strip()
        open_value = (
            results.find("td", attrs={"data-test": "OPEN-value"})
        ).text.strip()
        eps = (results.find("td", attrs={"data-test": "EPS_RATIO-value"})).text.strip()
        pe_ratio = (
            results.find("td", attrs={"data-test": "PE_RATIO-value"})
        ).text.strip()
        dividend_and_yield = (
            results.find("td", attrs={"data-test": "DIVIDEND_AND_YIELD-value"})
        ).text.strip()
        # stock_data=results.find_all("td", class_="Ta(end) Fw(600) Lh(14px)") << one way to get all the values more easily

        # Get the historical data for 1 month
        h_url = f"https://www.marketwatch.com/investing/stock/{symbol}/downloaddatapartial?startdate={month_ago}%2000:00:00&enddate={today}%2023:59:59&daterange=d30&frequency=p1d&csvdownload=true&downloadpartial=false&newdates=false"
        h_r = requests.get(h_url)

        # Format
        h_data = h_r.content.decode("utf-8")
        h = csv.reader(h_data.splitlines(), delimiter=",")
        history = list(h)

        # Create the worksheet
        ws = wb.create_sheet(f"{symbol}")

        # Write the columns (cells can be fixed)
        ws.cell(row=1, column=1, value="Corp Name")
        ws.cell(row=1, column=2, value="Previous Close")
        ws.cell(row=1, column=3, value="Open Value")
        ws.cell(row=1, column=4, value="EPS Ratio")
        ws.cell(row=1, column=5, value="PE Ratio")
        ws.cell(row=1, column=6, value="Dividend & Yield")
        ws.cell(row=1, column=7, value="Date")

        # Write the basic values to each sheet
        ws.cell(row=2, column=1, value=corp)
        ws.cell(row=2, column=2, value=prev_close)
        ws.cell(row=2, column=3, value=open_value)
        ws.cell(row=2, column=4, value=eps)
        ws.cell(row=2, column=5, value=pe_ratio)
        ws.cell(row=2, column=6, value=dividend_and_yield)
        ws.cell(row=2, column=7, value=today)

        # ------------Last 1 Month------------ #
        ws.append([])
        ws.append([])

        for row in history:
            ws.append(row)

wb.save(f"BasicReport-{date.today()}.xlsx")
print("created worksheets: " + str(wb.sheetnames))
